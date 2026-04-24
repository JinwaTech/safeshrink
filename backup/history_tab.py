"""
历史记录标签页
功能：显示文件处理历史记录
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QFrame, QMessageBox, QMenu, QScrollArea
)
from datetime import datetime
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction, QCursor, QColor
from history_manager import HistoryManager


class HistoryTab(QWidget):
    def __init__(self, shared_manager=None):
        super().__init__()
        # 使用共享的历史管理器，避免多实例数据不同步
        self.history_manager = shared_manager if shared_manager else HistoryManager()
        self.setup_ui()
        self.load_history()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 16, 0, 0)
        layout.setSpacing(16)
        
        # 标题区域
        title_frame = QFrame()
        title_layout = QHBoxLayout(title_frame)
        title_layout.setContentsMargins(20, 0, 20, 0)
        
        title = QLabel("📋 处理历史")
        title.setObjectName("pageTitle")
        
        # 统计信息
        self.stats_label = QLabel("共 0 条记录")
        self.stats_label.setObjectName("pageSubtitle")
        
        # 按钮
        btn_refresh = QPushButton("刷新")
        btn_refresh.clicked.connect(self.load_history)
        
        btn_export = QPushButton("导出报告")
        btn_export.setObjectName("secondary")
        btn_export.clicked.connect(self.export_report)
        
        btn_clear = QPushButton("清空历史")
        btn_clear.setObjectName("danger")
        btn_clear.clicked.connect(self.clear_history)
        
        title_layout.addWidget(title)
        title_layout.addWidget(self.stats_label)
        title_layout.addStretch()
        title_layout.addWidget(btn_export)
        title_layout.addWidget(btn_refresh)
        title_layout.addWidget(btn_clear)
        
        layout.addWidget(title_frame)
        
        # 可滚动内容区
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(16)
        
        # 历史记录表格
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["时间", "文件名", "操作", "处理前", "处理后", "效果"])
        
        # 设置列宽
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # 时间
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)        # 文件名
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # 操作
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # 处理前
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # 处理后
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # 效果
        self.table.horizontalHeader().setMinimumSectionSize(60)   # 通用最小宽度
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # 文件名占满剩余
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        # 样式由 theme_manager 统一设置，不在这里硬编码
        scroll_layout.addWidget(self.table)
        
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll, 1)
        
        # 底部提示
        bottom_frame = QFrame()
        bottom_layout = QHBoxLayout(bottom_frame)
        bottom_layout.setContentsMargins(20, 0, 20, 0)
        
        hint = QLabel("💡 提示：右键点击记录可以打开或删除")
        hint.setStyleSheet("color: #8b92a5; font-size: 12px;")
        
        bottom_layout.addWidget(hint)
        bottom_layout.addStretch()
        
        layout.addWidget(bottom_frame)
    
    def load_history(self):
        """加载历史记录"""
        history = self.history_manager.get_history(50)
        
        self.table.setRowCount(len(history))
        
        for row, record in enumerate(history):
            # 判断是否是批量处理
            is_batch = record.get('batch', False)
            
            # 时间
            time_item = QTableWidgetItem(record['time'])
            time_item.setData(Qt.ItemDataRole.UserRole, record['id'])
            self.table.setItem(row, 0, time_item)
            
            # 文件名
            name_item = QTableWidgetItem(record['file_name'])
            name_item.setData(Qt.ItemDataRole.UserRole, record)
            self.table.setItem(row, 1, name_item)
            
            # 操作
            action_name = self.history_manager.get_action_name(record['action'])
            action_item = QTableWidgetItem(action_name)
            self.table.setItem(row, 2, action_item)
            
            if is_batch:
                # 批量处理
                if record.get('action') == 'sanitize':
                    items = record.get('total_items', 0)
                    self.table.setItem(row, 3, QTableWidgetItem(f"检测到 {items} 处"))
                    self.table.setItem(row, 4, QTableWidgetItem(f"剩余 0 处"))
                    effect_text = f"脱敏 {items}处" if items > 0 else "-"
                    effect_item = QTableWidgetItem(effect_text)
                    if items > 0:
                        effect_item.setForeground(QColor('#38a169'))
                    self.table.setItem(row, 5, effect_item)
                else:
                    # 批量减肥/SSD：尝试显示 Token，否则显示文件大小
                    orig_tok = record.get('original_tokens')
                    new_tok = record.get('new_tokens')
                    if orig_tok and new_tok and orig_tok > 0:
                        saved_tok = orig_tok - new_tok
                        self.table.setItem(row, 3, QTableWidgetItem(f"~{orig_tok:,}"))
                        self.table.setItem(row, 4, QTableWidgetItem(f"~{new_tok:,}"))
                        if saved_tok > 0:
                            tok_text = f"↓ {saved_tok:,} token"
                        else:
                            tok_text = "-"
                        tok_item = QTableWidgetItem(tok_text)
                        if saved_tok > 0:
                            tok_item.setForeground(QColor('#38a169'))
                        self.table.setItem(row, 5, tok_item)
                    else:
                        orig_size = self.history_manager.format_size(record.get('original_size', 0))
                        new_size = self.history_manager.format_size(record.get('new_size', 0))
                        self.table.setItem(row, 3, QTableWidgetItem(orig_size))
                        self.table.setItem(row, 4, QTableWidgetItem(new_size))
                        saved = record.get('saved_percent', 0)
                        saved_size = record.get('saved_size', 0)
                        if saved > 0:
                            effect_text = f"↓ {saved}% (省{self.history_manager.format_size(saved_size)})"
                        else:
                            effect_text = "-"
                        effect_item = QTableWidgetItem(effect_text)
                        if saved > 0:
                            effect_item.setForeground(QColor('#38a169'))
                        self.table.setItem(row, 5, effect_item)
            else:
                # 普通文件：尝试显示 Token（SSD），否则显示文件大小
                orig_tok = record.get('original_tokens')
                new_tok = record.get('new_tokens')
                # estimate_tokens 返回 dict，需要提取 total
                if isinstance(orig_tok, dict):
                    orig_tok = orig_tok.get('total', 0)
                if isinstance(new_tok, dict):
                    new_tok = new_tok.get('total', 0)
                if orig_tok and new_tok and orig_tok > 0:
                    saved_tok = orig_tok - new_tok
                    self.table.setItem(row, 3, QTableWidgetItem(f"~{orig_tok:,}"))
                    self.table.setItem(row, 4, QTableWidgetItem(f"~{new_tok:,}"))
                    if saved_tok > 0:
                        tok_text = f"↓ {saved_tok:,} token"
                    else:
                        tok_text = "-"
                    tok_item = QTableWidgetItem(tok_text)
                    if saved_tok > 0:
                        tok_item.setForeground(QColor('#38a169'))
                    self.table.setItem(row, 5, tok_item)
                else:
                    # 脱敏或无 token 数据：显示文件大小
                    orig_size = self.history_manager.format_size(record['original_size'])
                    self.table.setItem(row, 3, QTableWidgetItem(orig_size))
                    new_size = self.history_manager.format_size(record['new_size'])
                    self.table.setItem(row, 4, QTableWidgetItem(new_size))
                    if record.get('action') == 'sanitize':
                        total_items = record.get('total_items', 0)
                        if total_items > 0:
                            saved_text = f"脱敏 {total_items}处"
                        else:
                            saved_text = "-"
                        saved_item = QTableWidgetItem(saved_text)
                        if total_items > 0:
                            saved_item.setForeground(QColor('#38a169'))
                    else:
                        saved = record.get('saved_size', 0)
                        if saved > 0:
                            pct = record.get('saved_percent', 0)
                            saved_text = f"↓ {self.history_manager.format_size(saved)}"
                            if pct > 0:
                                saved_text += f" ({pct}%)"
                        else:
                            saved_text = "-"
                        saved_item = QTableWidgetItem(saved_text)
                        if saved > 0:
                            saved_item.setForeground(QColor('#38a169'))
                    self.table.setItem(row, 5, saved_item)
        self.table.resizeColumnsToContents()
        self.stats_label.setText(f"共 {len(history)} 条记录")
    
    def show_context_menu(self, position):
        """显示右键菜单"""
        row = self.table.currentRow()
        if row < 0:
            return
        
        record = self.table.item(row, 1).data(Qt.ItemDataRole.UserRole)
        if not record:
            return
        
        menu = QMenu(self)
        
        # 判断是否是批量处理
        is_batch = record.get('batch', False)
        
        # 打开输出文件夹（批量处理或普通文件）
        if is_batch:
            # 批量处理：打开 output 文件夹
            action_folder = QAction("打开输出文件夹", self)
            action_folder.triggered.connect(lambda: self.open_batch_folder(record))
            menu.addAction(action_folder)
        elif record.get('output_path'):
            # 普通文件：打开文件和文件夹
            action_open = QAction("打开输出文件", self)
            action_open.triggered.connect(lambda: self.open_file(record['output_path']))
            menu.addAction(action_open)
            
            action_folder = QAction("打开输出文件夹", self)
            action_folder.triggered.connect(lambda: self.open_folder(record['output_path']))
            menu.addAction(action_folder)
        elif record.get('file_path'):
            # 有原文件路径，打开原文件所在文件夹
            action_folder = QAction("打开原文件夹", self)
            action_folder.triggered.connect(lambda: self.open_folder(record['file_path']))
            menu.addAction(action_folder)
        else:
            # 无路径信息
            action_none = QAction("无可用路径", self)
            action_none.setEnabled(False)
            menu.addAction(action_none)
        
        menu.addSeparator()
        
        # 删除
        action_delete = QAction("删除此记录", self)
        action_delete.triggered.connect(lambda: self.delete_record(record['id']))
        menu.addAction(action_delete)
        
        menu.exec(QCursor.pos())
    
    def open_batch_folder(self, record):
        """打开批量处理的输出文件夹"""
        import os
        from pathlib import Path
        
        # 尝试从 file_path 推断 output 文件夹
        if record.get('file_path'):
            folder = Path(record['file_path']).parent / 'output'
        else:
            # 使用默认位置
            folder = Path.home() / 'Desktop' / 'output'
        
        if folder.exists():
            os.startfile(str(folder))
        else:
            QMessageBox.warning(self, "提示", f"输出文件夹不存在:\n{folder}")
    
    def open_file(self, file_path):
        """打开文件"""
        import os
        if file_path and os.path.exists(file_path):
            os.startfile(file_path)
        else:
            QMessageBox.warning(self, "提示", "文件不存在或已被移动")
    
    def open_folder(self, file_path):
        """打开文件夹"""
        import os
        if file_path and os.path.exists(file_path):
            folder = os.path.dirname(file_path)
            os.startfile(folder)
        else:
            QMessageBox.warning(self, "提示", "文件夹不存在或已被移动")
    
    def delete_record(self, record_id):
        """删除记录"""
        reply = QMessageBox.question(
            self, "确认", "确定要删除这条记录吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.history_manager.delete_record(record_id)
            self.load_history()
    
    def clear_history(self):
        """清空历史"""
        reply = QMessageBox.question(
            self, "确认", "确定要清空所有历史记录吗？\n此操作不可恢复。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.history_manager.clear_history()
            self.load_history()
    
    def export_report(self):
        """导出处理报告"""
        history = self.history_manager.get_history(500)
        if not history:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示", "没有可导出的历史记录。")
            return

        from PyQt6.QtWidgets import QFileDialog, QInputDialog
        
        # 选择格式
        fmt, ok = QInputDialog.getItem(
            self, "导出格式", "请选择导出格式：",
            ["CSV 文件 (*.csv)", "Excel 文件 (*.xlsx)"], 0, False
        )
        if not ok:
            return
        
        ext = '.csv' if 'CSV' in fmt else '.xlsx'
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出报告",
            f"SafeShrink_报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}",
            fmt
        )
        
        if not file_path:
            return

        try:
            # 表头
            headers = ['时间', '文件名', '操作', '处理前', '处理后', '效果']

            # 处理前/后格式化（SSD 显示 Token，其他显示文件大小）
            def get_size_cols(item):
                orig_tok = item.get('original_tokens')
                new_tok = item.get('new_tokens')
                if orig_tok and new_tok and orig_tok > 0:
                    return f"~{orig_tok:,}", f"~{new_tok:,}"
                orig = item.get('original_size', 0) or 0
                new_sz = item.get('new_size', 0) or 0
                return (self._format_size(orig) if orig > 0 else '-'), (self._format_size(new_sz) if new_sz > 0 else '-')

            # 效果计算
            def get_effect(item):
                action = item.get('action', 'slim')
                orig_tok = item.get('original_tokens')
                new_tok = item.get('new_tokens')
                if orig_tok and new_tok and orig_tok > 0:
                    saved_tok = orig_tok - new_tok
                    if saved_tok > 0:
                        return f"↓ {saved_tok:,} token"
                    return '-'
                if action == 'slim':
                    saved = item.get('saved_size', 0) or 0
                    orig = item.get('original_size', 0) or 0
                    if saved > 0 and orig > 0:
                        pct = round((saved / orig) * 100, 1)
                        return f"↓ {self._format_size(saved)} ({pct}%)"
                    return '-'
                else:
                    count = item.get('items_found', item.get('success_count', '-'))
                    if count != '-':
                        return f"脱敏 {count} 处"
                    return '-'
            
            if ext == '.csv':
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    for item in history:
                        orig_sz_col, new_sz_col = get_size_cols(item)
                        writer.writerow([
                            item.get("time", ""),
                            item.get("file_name", ""),
                            action_tag,
                            orig_sz_col,
                            new_sz_col,
                            get_effect(item),
                        ])
            else:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "处理报告"
                
                # 表头样式
                hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                hfont = Font(bold=True, color='FFFFFF')
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = hf
                    cell.font = hfont
                    cell.alignment = Alignment(horizontal='center')
                
                # 数据
                for row_idx, item in enumerate(history, 2):
                    action_tag = "减肥" if item.get("action") == "slim" else "脱敏"
                    orig_sz_col, new_sz_col = get_size_cols(item)
                    ws.cell(row=row_idx, column=1, value=item.get("time", ""))
                    ws.cell(row=row_idx, column=2, value=item.get("file_name", ""))
                    ws.cell(row=row_idx, column=3, value=action_tag)
                    ws.cell(row=row_idx, column=4, value=orig_sz_col)
                    ws.cell(row=row_idx, column=5, value=new_sz_col)
                    ws.cell(row=row_idx, column=6, value=get_effect(item))


                # 列宽
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 18
                ws.column_dimensions['F'].width = 22
                ws.column_dimensions['G'].width = 30
                
                wb.save(file_path)
            
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "完成", f"报告已导出到：\n{file_path}")
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "错误", f"导出失败：{e}")


    def _format_size(self, size):
        """格式化文件大小"""
        try:
            size = int(size)
            if size < 1024:
                return f"{size} B"
            elif size < 1024 * 1024:
                return f"{size/1024:.1f} KB"
            else:
                return f"{size/(1024*1024):.2f} MB"
        except:
            return "0 B"

    def add_record(self, file_name, file_path, action, original_size, new_size, output_path=None):
        """添加记录（供其他标签页调用）"""
        return self.history_manager.add_record(
            file_name, file_path, action, original_size, new_size, output_path
        )
    
    def add_batch_record(self, total_files, success_count, total_saved, action='slim'):
        """添加批量处理记录"""
        return self.history_manager.add_batch_record(
            total_files, success_count, total_saved, action
        )
    
    def update_language(self, lang):
        """更新语言"""
        from translations import get_translation
        _ = lambda t: get_translation(t, lang)
        
        # 更新按钮
        if hasattr(self, 'btn_clear'):
            self.btn_clear.setText(_('清空历史'))
        if hasattr(self, 'btn_export'):
            self.btn_export.setText(_('导出历史'))
        
        # 更新表格标题
        if hasattr(self, 'table'):
            self.table.setHorizontalHeaderLabels([
                _('时间'), _('文件名'), _('操作'), _('原大小'), _('新大小'), _('状态'), _('详情')
            ])
