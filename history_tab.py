"""

历史记录标签页

功能：显示文件处理历史记录

"""



from PySide6.QtWidgets import (

    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 

    QPushButton, QTableWidget, QTableWidgetItem,

    QHeaderView, QFrame, QMessageBox, QMenu, QScrollArea,

    QDateEdit, QComboBox, QTableView, QLineEdit, QAbstractSpinBox, QToolButton

)

from PySide6.QtCore import QDate, QSize
from PySide6.QtWidgets import QSizePolicy



from datetime import datetime

from PySide6.QtCore import Qt

from PySide6.QtGui import QAction, QCursor, QColor

from history_manager import HistoryManager
from theme_manager import ThemeManager





class HistoryTab(QWidget):

    def __init__(self, shared_manager=None):

        super().__init__()

        # 使用共享的历史管理器，避免多实例数据不同步

        self.history_manager = shared_manager if shared_manager else HistoryManager()

        self.setup_ui()

        self.load_history()

    

    def setup_ui(self):

        layout = QVBoxLayout(self)

        layout.setContentsMargins(0, 16, 0, 0)

        layout.setSpacing(16)

        

        # 标题区域

        title_frame = QFrame()

        title_layout = QHBoxLayout(title_frame)

        title_layout.setContentsMargins(20, 0, 20, 0)

        

        title = QLabel("📋 处理历史")

        title.setObjectName("pageTitle")

        

        # 统计信息

        self.stats_label = QLabel("共 0 条记录")

        self.stats_label.setObjectName("pageSubtitle")

        

        # 按钮

        btn_refresh = QPushButton("刷新")

        btn_refresh.clicked.connect(self.load_history)

        

        btn_export = QPushButton("导出报告")

        btn_export.setObjectName("secondary")

        btn_export.clicked.connect(self.export_report)

        

        btn_clear = QPushButton("清空历史")

        btn_clear.setObjectName("danger")

        btn_clear.clicked.connect(self.clear_history)

        

        title_layout.addWidget(title)

        title_layout.addWidget(self.stats_label)

        title_layout.addStretch()

        title_layout.addWidget(btn_export)

        title_layout.addWidget(btn_refresh)

        title_layout.addWidget(btn_clear)

        

        layout.addWidget(title_frame)

        

        # 可滚动内容区

        scroll = QScrollArea()

        scroll.setWidgetResizable(True)

        scroll.setFrameShape(QFrame.Shape.NoFrame)

        

        scroll_content = QWidget()

        scroll_layout = QVBoxLayout(scroll_content)

        scroll_layout.setContentsMargins(0, 0, 0, 0)

        scroll_layout.setSpacing(16)

        

        # 历史记录表格

        # -- Filter Bar --

        filter_frame = QFrame()

        filter_layout = QHBoxLayout(filter_frame)

        filter_layout.setContentsMargins(20, 8, 20, 8)

        filter_layout.setSpacing(12)

        filter_layout.addWidget(QLabel("筛选日期："))

        self.filter_date_from = QDateEdit()

        self.filter_date_from.setCalendarPopup(True)

        self.filter_date_from.setDateRange(QDate(2026, 1, 1), QDate.currentDate())

        self.filter_date_from.setDate(QDate(2026, 1, 1))

        self.filter_date_from.setFixedWidth(130)

        filter_layout.addWidget(self.filter_date_from)

        filter_layout.addWidget(QLabel("—"))

        self.filter_date_to = QDateEdit()

        self.filter_date_to.setCalendarPopup(True)

        self.filter_date_to.setDateRange(QDate(2026, 1, 1), QDate.currentDate())

        self.filter_date_to.setDate(QDate.currentDate())

        self.filter_date_to.setFixedWidth(130)

        filter_layout.addWidget(self.filter_date_to)

        self.filter_action = QComboBox()

        self.filter_action.addItems(["全部", "文件减肠", "文件脱敏", "批量减肠", "批量脱敏"])

        self.filter_action.currentIndexChanged.connect(self._on_filter_changed)

        filter_layout.addWidget(self.filter_action)

        self.filter_name = QLineEdit()

        self.filter_name.setPlaceholderText("文件名筛选")

        self.filter_name.textChanged.connect(self._on_filter_changed)

        self.filter_name.setFixedWidth(160)

        filter_layout.addWidget(self.filter_name)

        btn_reset = QPushButton("重置")

        btn_reset.clicked.connect(self._reset_filter)

        filter_layout.addWidget(btn_reset)

        filter_layout.addStretch()

        layout.addWidget(filter_frame)

        # 日历弹窗事件过滤（延迟安装，见 eventFilter）
        self._calendar_filtered = {self.filter_date_from: False,
                                    self.filter_date_to: False}

        # ── 为日历弹窗安装 eventFilter ──────────────────────────
        # calendarWidget() 在弹窗首次打开前返回 None，所以先装在 date_edit 上，
        # 等弹窗 Show 时再抓到真正的日历 widget 并装上
        self.filter_date_from.installEventFilter(self)
        self.filter_date_to.installEventFilter(self)



        self.table = QTableWidget()

        self.table.setColumnCount(6)

        self.table.setHorizontalHeaderLabels(["时间", "文件名", "操作", "处理前", "处理后", "效果"])

        

        # 设置列宽

        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # 时间

        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)        # 文件名

        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # 操作

        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # 处理前

        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # 处理后

        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # 效果

        self.table.horizontalHeader().setMinimumSectionSize(60)   # 通用最小宽度

        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # 文件名占满剩余

        self.table.setAlternatingRowColors(True)

        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)

        self.table.customContextMenuRequested.connect(self.show_context_menu)

        # 样式由 theme_manager 统一设置，不在这里硬编码

        scroll_layout.addWidget(self.table)

        

        scroll.setWidget(scroll_content)

        layout.addWidget(scroll, 1)

        

        # 底部提示

        bottom_frame = QFrame()

        bottom_layout = QHBoxLayout(bottom_frame)

        bottom_layout.setContentsMargins(20, 0, 20, 0)

        

        hint = QLabel("💡 提示：右键点击记录可以打开或删除")

        hint.setStyleSheet("color: #8b92a5; font-size: 12px;")

        

        bottom_layout.addWidget(hint)

        bottom_layout.addStretch()

        

        layout.addWidget(bottom_frame)

    

    def load_history(self):

        """加载历史记录"""

        history = self.history_manager.get_history(50)

        

        # filter

        history = self._filter_history(history)



        self.table.setRowCount(len(history))

        

        for row, record in enumerate(history):

            # 判断是否是批量处理

            is_batch = record.get('batch', False)

            

            # 时间

            time_item = QTableWidgetItem(record['time'])

            time_item.setData(Qt.ItemDataRole.UserRole, record['id'])

            self.table.setItem(row, 0, time_item)

            

            # 文件名

            name_item = QTableWidgetItem(record['file_name'])

            name_item.setData(Qt.ItemDataRole.UserRole, record)

            self.table.setItem(row, 1, name_item)

            

            # 操作

            action_name = self.history_manager.get_action_name(record['action'])

            action_item = QTableWidgetItem(action_name)

            self.table.setItem(row, 2, action_item)

            

            if is_batch:

                # 批量处理

                if record.get('action') == 'sanitize':

                    items = record.get('total_items', 0)

                    self.table.setItem(row, 3, QTableWidgetItem(f"检测到 {items} 处"))

                    self.table.setItem(row, 4, QTableWidgetItem(f"剩余 0 处"))

                    effect_text = f"脱敏 {items}处" if items > 0 else "-"

                    effect_item = QTableWidgetItem(effect_text)

                    if items > 0:

                        effect_item.setForeground(QColor('#38a169'))

                    self.table.setItem(row, 5, effect_item)

                else:

                    # 批量减肥/SSD：尝试显示 Token，否则显示文件大小

                    orig_tok = record.get('original_tokens')

                    new_tok = record.get('new_tokens')

                    if orig_tok and new_tok and orig_tok > 0:

                        saved_tok = orig_tok - new_tok

                        self.table.setItem(row, 3, QTableWidgetItem(f"~{orig_tok:,}"))

                        self.table.setItem(row, 4, QTableWidgetItem(f"~{new_tok:,}"))

                        if saved_tok > 0:

                            tok_text = f"↓ {saved_tok:,} token"

                        else:

                            tok_text = "-"

                        tok_item = QTableWidgetItem(tok_text)

                        if saved_tok > 0:

                            tok_item.setForeground(QColor('#38a169'))

                        self.table.setItem(row, 5, tok_item)

                    else:

                        orig_size = self.history_manager.format_size(record.get('original_size', 0))

                        new_size = self.history_manager.format_size(record.get('new_size', 0))

                        self.table.setItem(row, 3, QTableWidgetItem(orig_size))

                        self.table.setItem(row, 4, QTableWidgetItem(new_size))

                        saved = record.get('saved_percent', 0)

                        saved_size = record.get('saved_size', 0)

                        if saved > 0:

                            effect_text = f"↓ {saved}% (省{self.history_manager.format_size(saved_size)})"

                        else:

                            effect_text = "-"

                        effect_item = QTableWidgetItem(effect_text)

                        if saved > 0:

                            effect_item.setForeground(QColor('#38a169'))

                        self.table.setItem(row, 5, effect_item)

            else:

                # 普通文件：尝试显示 Token（SSD），否则显示文件大小

                orig_tok = record.get('original_tokens')

                new_tok = record.get('new_tokens')

                # estimate_tokens 返回 dict，需要提取 total

                if isinstance(orig_tok, dict):

                    orig_tok = orig_tok.get('total', 0)

                if isinstance(new_tok, dict):

                    new_tok = new_tok.get('total', 0)

                if orig_tok and new_tok and orig_tok > 0:

                    saved_tok = orig_tok - new_tok

                    self.table.setItem(row, 3, QTableWidgetItem(f"~{orig_tok:,}"))

                    self.table.setItem(row, 4, QTableWidgetItem(f"~{new_tok:,}"))

                    if saved_tok > 0:

                        tok_text = f"↓ {saved_tok:,} token"

                    else:

                        tok_text = "-"

                    tok_item = QTableWidgetItem(tok_text)

                    if saved_tok > 0:

                        tok_item.setForeground(QColor('#38a169'))

                    self.table.setItem(row, 5, tok_item)

                else:

                    # 脱敏或无 token 数据：显示文件大小

                    orig_size = self.history_manager.format_size(record['original_size'])

                    self.table.setItem(row, 3, QTableWidgetItem(orig_size))

                    new_size = self.history_manager.format_size(record['new_size'])

                    self.table.setItem(row, 4, QTableWidgetItem(new_size))

                    if record.get('action') == 'sanitize':

                        total_items = record.get('total_items', 0)

                        if total_items > 0:

                            saved_text = f"脱敏 {total_items}处"

                        else:

                            saved_text = "-"

                        saved_item = QTableWidgetItem(saved_text)

                        if total_items > 0:

                            saved_item.setForeground(QColor('#38a169'))

                    else:

                        saved = record.get('saved_size', 0)

                        if saved > 0:

                            pct = record.get('saved_percent', 0)

                            saved_text = f"↓ {self.history_manager.format_size(saved)}"

                            if pct > 0:

                                saved_text += f" ({pct}%)"

                        else:

                            saved_text = "-"

                        saved_item = QTableWidgetItem(saved_text)

                        if saved > 0:

                            saved_item.setForeground(QColor('#38a169'))

                    self.table.setItem(row, 5, saved_item)

        self.table.resizeColumnsToContents()

        self.stats_label.setText(f"共 {len(history)} 条记录")

    

    def eventFilter(self, obj, event):
        from PySide6.QtCore import QEvent
        if event.type() == QEvent.Type.Show:
            for de in [self.filter_date_from, self.filter_date_to]:
                if obj != de:
                    continue
                cw = de.calendarWidget()
                if cw is None:
                    continue
                ThemeManager.apply_calendar_style(cw, 'dark')
                cw.setMinimumWidth(320)
                # 导航栏间隙修复：找到导航栏和按钮，统一视觉
                from PySide6.QtWidgets import QWidget as QW, QApplication
                real_cal = None  # 真正的日历弹窗窗口
                nav_bar = None   # 导航栏
                for child in cw.findChildren(QW):
                    name = child.objectName()
                    if name == 'qt_calendar_navigationbar':
                        nav_bar = child
                        # 导航栏整体：统一背景色，spacing=0
                        child.setStyleSheet('''
                            QWidget#qt_calendar_navigationbar {
                                background-color: #313244;
                                border-bottom: 1px solid #45475a;
                            }
                        ''')
                        lay = child.layout()
                        if lay:
                            lay.setSpacing(2)
                            lay.setContentsMargins(6, 4, 6, 4)
                            # 收紧 spacer[1] 和 spacer[3]，spacer[5] 自动变大给 +/- 让位
                            for i in range(lay.count()):
                                item = lay.itemAt(i)
                                if item.spacerItem():
                                    if i in (1, 3):
                                        item.spacerItem().changeSize(5, 30)
                            # 年份按钮设最小宽度，防止显示 "202" 而非 "2026"
                            for qbtn in cw.findChildren(QW):
                                if qbtn.objectName() == 'qt_calendar_yearbutton':
                                    qbtn.setMinimumWidth(65)
                                    break
                    elif name in ('qt_calendar_prevmonth', 'qt_calendar_nextmonth'):
                        child.setStyleSheet('QToolButton{color:#cdd6f4!important;background:transparent;border:none}QToolButton:hover{background-color:rgba(137,180,250,0.25);color:#89b4fa!important}')
                    elif name in ('qt_calendar_monthbutton', 'qt_calendar_yearbutton'):
                        # 月份/年份：无边框，与导航栏同色背景
                        btn_style = '''
                            QToolButton {
                                background-color: #313244;
                                color: #cdd6f4;
                                border: none;
                                border-radius: 0;
                                padding: 4px 8px;
                                margin: 0;
                            }
                            QToolButton:hover {
                                background-color: rgba(137,180,250,0.25);
                            }
                            QToolButton::menu-indicator { image: none; }
                        '''
                        child.setStyleSheet(btn_style)
                # 下拉菜单和年份选择器样式
                cw.setStyleSheet(cw.styleSheet() + '''
                    QCalendarWidget QMenu {
                        background-color: #313244;
                        border: 1px solid #45475a;
                    }
                    QCalendarWidget QMenu::item {
                        background-color: #313244;
                        color: #cdd6f4;
                        padding: 4px 20px;
                    }
                    QCalendarWidget QMenu::item:selected {
                        background-color: #89b4fa;
                        color: #1e1e2e;
                    }
                    QCalendarWidget QSpinBox {
                        background-color: #313244;
                        color: #cdd6f4;
                        border: 1px solid #45475a;
                        padding: 2px 20px 2px 4px;
                        font-size: 13px;
                    }
                    QCalendarWidget QSpinBox:focus {
                        border: 1px solid #89b4fa;
                    }
                ''')
                # 年份选择器 QSpinBox → 只处理年份，添加与导航箭头同款的 +/- QToolButton
                from PySide6.QtWidgets import QSpinBox, QLabel, QToolButton
                for sb in cw.findChildren(QSpinBox):
                    if hasattr(sb, '_safeshrink_custom_buttons'):
                        continue
                    if not (1900 <= sb.value() <= 2100):
                        continue
                    # 隐藏 spinbox 原生箭头
                    sb.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
                    # 找导航栏
                    nav_bar = None
                    for child in cw.findChildren(QW):
                        if child.objectName() == 'qt_calendar_navigationbar':
                            nav_bar = child
                            break
                    if not nav_bar:
                        continue
                    # 与 ◀/▶ 同款：transparent 背景 + #cdd6f4 文字
                    arrow_style = '''
                        QToolButton {
                            background: transparent;
                            color: #cdd6f4;
                            border: none;
                            padding: 2px 8px;
                            font-size: 13px;
                        }
                        QToolButton:hover {
                            background-color: rgba(137,180,250,0.25);
                            color: #89b4fa;
                        }
                        QToolButton:pressed {
                            background-color: #89b4fa;
                            color: #1e1e2e;
                        }
                    '''
                    nav_btn_style = 'QToolButton{color:#cdd6f4!important;background:transparent;border:none}QToolButton:hover{background-color:rgba(137,180,250,0.25)}'
                    btn_up = QToolButton(nav_bar)
                    btn_down = QToolButton(nav_bar)
                    btn_up.setText('+')
                    btn_down.setText('−')
                    for btn in (btn_up, btn_down):
                        btn.setStyleSheet(nav_btn_style)
                        btn.setCursor(Qt.CursorShape.PointingHandCursor)
                    sb_ref = sb
                    btn_up.clicked.connect(lambda _, s=sb_ref: s.setValue(s.value() + 1))
                    btn_down.clicked.connect(lambda _, s=sb_ref: s.setValue(s.value() - 1))
                    # spacer[1]/[3] 收紧后，spacer[5] 自动变大，+/− 用 insertWidget 插入导航栏 layout
                    nav_lay = nav_bar.layout()
                    if nav_lay:
                        # 找到年份按钮在 layout 中的位置，紧贴其后插入 +/-
                        year_idx = -1
                        for i in range(nav_lay.count()):
                            w = nav_lay.itemAt(i).widget()
                            if w and w.objectName() == 'qt_calendar_yearbutton':
                                year_idx = i
                                break
                        if year_idx >= 0:
                            nav_lay.insertWidget(year_idx + 1, btn_up)
                            nav_lay.insertWidget(year_idx + 2, btn_down)
                        else:
                            # fallback：插到末尾（▶ 前面）
                            nav_lay.insertWidget(nav_lay.count() - 1, btn_up)
                            nav_lay.insertWidget(nav_lay.count() - 1, btn_down)
                    else:
                        # 无 layout 时 fallback 到 move 定位
                        btn_up.move(155, 2)
                        btn_down.move(185, 2)
                        btn_up.show()
                        btn_down.show()
                    sb._safeshrink_custom_buttons = (btn_up, btn_down)
                # 星期头高度压到 20px
                day_names = {chr(0x4e00), chr(0x4e8c), chr(0x4e09), chr(0x56db), chr(0x4e94), chr(0x516d), chr(0x65e5)}
                day_en = {'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'}
                for lbl in cw.findChildren(QLabel):
                    t = lbl.text().strip()
                    if t in day_names or t in day_en:
                        lbl.setFixedHeight(20)
                break
        return super().eventFilter(obj, event)



    def _on_filter_changed(self):

        """任意筛选标签变化时重新加载"""

        self.load_history()



    def _reset_filter(self):

        """重置所有筛选条件"""

        self.filter_name.clear()

        self.filter_action.setCurrentIndex(0)

        self.filter_date_from.setDate(QDate(2026, 1, 1))

        self.filter_date_to.setDate(QDate.currentDate())



    def _date_to_str(self, qdate):

        """将 QDate 转为 YYYY-MM-DD；2026-01-01 返回 None"""

        if qdate.year() == 2026 and qdate.month() == 1 and qdate.day() == 1:

            return None

        return qdate.toString("yyyy-MM-dd")



    def _filter_history(self, history):

        """根据筛选条件过滤历史"""

        action_idx = self.filter_action.currentIndex()

        name_pattern = self.filter_name.text().strip().lower()

        date_from = self._date_to_str(self.filter_date_from.date())

        date_to   = self._date_to_str(self.filter_date_to.date())

        if action_idx == 0 and not name_pattern and not date_from and not date_to:

            return history

        filtered = []

        for record in history:

            if action_idx > 0:

                is_batch = record.get("batch", False)

                if action_idx == 3 and not is_batch: continue

                if action_idx == 4 and not is_batch: continue

                if action_idx in (1, 2) and is_batch: continue

                rec_action = record.get("action", "")

                expected = {1: "slim", 2: "sanitize", 3: "slim_batch", 4: "sanitize_batch"}

                if expected.get(action_idx) and rec_action != expected[action_idx]: continue

            if name_pattern:

                if name_pattern not in record.get("file_name", "").lower(): continue

            rec_time = record.get("time", "")

            if rec_time:

                try:

                    rec_date = rec_time[:10]

                    if date_from and rec_date < date_from: continue

                    if date_to and rec_date > date_to: continue

                except Exception: pass

            filtered.append(record)

        return filtered



    def show_context_menu(self, position):

        """显示右键菜单"""

        row = self.table.currentRow()

        if row < 0:

            return

        

        record = self.table.item(row, 1).data(Qt.ItemDataRole.UserRole)

        if not record:

            return

        

        menu = QMenu(self)

        

        # 判断是否是批量处理

        is_batch = record.get('batch', False)

        

        # 打开输出文件夹（批量处理或普通文件）

        if is_batch:

            # 批量处理：打开 output 文件夹

            action_folder = QAction("打开输出文件夹", self)

            action_folder.triggered.connect(lambda: self.open_batch_folder(record))

            menu.addAction(action_folder)

        elif record.get('output_path'):

            # 普通文件：打开文件和文件夹

            action_open = QAction("打开输出文件", self)

            action_open.triggered.connect(lambda: self.open_file(record['output_path']))

            menu.addAction(action_open)

            

            action_folder = QAction("打开输出文件夹", self)

            action_folder.triggered.connect(lambda: self.open_folder(record['output_path']))

            menu.addAction(action_folder)

        elif record.get('file_path'):

            # 有原文件路径，打开原文件所在文件夹

            action_folder = QAction("打开原文件夹", self)

            action_folder.triggered.connect(lambda: self.open_folder(record['file_path']))

            menu.addAction(action_folder)

        else:

            # 无路径信息

            action_none = QAction("无可用路径", self)

            action_none.setEnabled(False)

            menu.addAction(action_none)

        

        menu.addSeparator()

        

        # 删除

        action_delete = QAction("删除此记录", self)

        action_delete.triggered.connect(lambda: self.delete_record(record['id']))

        menu.addAction(action_delete)

        

        menu.exec(QCursor.pos())

    

    def open_batch_folder(self, record):

        """打开批量处理的输出文件夹"""

        import os

        from pathlib import Path

        

        # 尝试从 file_path 推断 output 文件夹

        if record.get('file_path'):

            folder = Path(record['file_path']).parent / 'output'

        else:

            # 使用默认位置

            folder = Path.home() / 'Desktop' / 'output'

        

        if folder.exists():

            os.startfile(str(folder))

        else:

            QMessageBox.warning(self, "提示", f"输出文件夹不存在:\n{folder}")

    

    def open_file(self, file_path):

        """打开文件"""

        import os

        if file_path and os.path.exists(file_path):

            os.startfile(file_path)

        else:

            QMessageBox.warning(self, "提示", "文件不存在或已被移动")

    

    def open_folder(self, file_path):

        """打开文件夹"""

        import os

        if file_path and os.path.exists(file_path):

            folder = os.path.dirname(file_path)

            os.startfile(folder)

        else:

            QMessageBox.warning(self, "提示", "文件夹不存在或已被移动")

    

    def delete_record(self, record_id):

        """删除记录"""

        reply = QMessageBox.question(

            self, "确认", "确定要删除这条记录吗？",

            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No

        )

        

        if reply == QMessageBox.StandardButton.Yes:

            self.history_manager.delete_record(record_id)

            self.load_history()

    

    def clear_history(self):

        """清空历史"""

        reply = QMessageBox.question(

            self, "确认", "确定要清空所有历史记录吗？\n此操作不可恢复。",

            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No

        )

        

        if reply == QMessageBox.StandardButton.Yes:

            self.history_manager.clear_history()

            self.load_history()

    

    def export_report(self):

        """导出处理报告"""

        history = self.history_manager.get_history(500)

        if not history:

            from PySide6.QtWidgets import QMessageBox

            QMessageBox.information(self, "提示", "没有可导出的历史记录。")

            return



        from PySide6.QtWidgets import QFileDialog, QInputDialog

        

        # 选择格式

        fmt, ok = QInputDialog.getItem(

            self, "导出格式", "请选择导出格式：",

            ["CSV 文件 (*.csv)", "Excel 文件 (*.xlsx)"], 0, False

        )

        if not ok:

            return

        

        ext = '.csv' if 'CSV' in fmt else '.xlsx'

        file_path, _ = QFileDialog.getSaveFileName(

            self, "导出报告",

            f"SafeShrink_报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}",

            fmt

        )

        

        if not file_path:

            return



        try:

            # 表头

            headers = ['时间', '文件名', '操作', '处理前', '处理后', '效果']



            # 处理前/后格式化（SSD 显示 Token，其他显示文件大小）

            def get_size_cols(item):

                orig_tok = item.get('original_tokens')

                new_tok = item.get('new_tokens')

                if orig_tok and new_tok and orig_tok > 0:

                    return f"~{orig_tok:,}", f"~{new_tok:,}"

                orig = item.get('original_size', 0) or 0

                new_sz = item.get('new_size', 0) or 0

                return (self._format_size(orig) if orig > 0 else '-'), (self._format_size(new_sz) if new_sz > 0 else '-')



            # 效果计算

            def get_effect(item):

                action = item.get('action', 'slim')

                orig_tok = item.get('original_tokens')

                new_tok = item.get('new_tokens')

                if orig_tok and new_tok and orig_tok > 0:

                    saved_tok = orig_tok - new_tok

                    if saved_tok > 0:

                        return f"↓ {saved_tok:,} token"

                    return '-'

                if action == 'slim':

                    saved = item.get('saved_size', 0) or 0

                    orig = item.get('original_size', 0) or 0

                    if saved > 0 and orig > 0:

                        pct = round((saved / orig) * 100, 1)

                        return f"↓ {self._format_size(saved)} ({pct}%)"

                    return '-'

                else:

                    count = item.get('items_found', item.get('success_count', '-'))

                    if count != '-':

                        return f"脱敏 {count} 处"

                    return '-'

            

            if ext == '.csv':

                import csv

                with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:

                    writer = csv.writer(csvfile)

                    writer.writerow(headers)

                    for item in history:

                        orig_sz_col, new_sz_col = get_size_cols(item)

                        writer.writerow([

                            item.get("time", ""),

                            item.get("file_name", ""),

                            action_tag,

                            orig_sz_col,

                            new_sz_col,

                            get_effect(item),

                        ])

            else:

                import openpyxl

                from openpyxl.styles import Font, Alignment, PatternFill

                

                wb = openpyxl.Workbook()

                ws = wb.active

                ws.title = "处理报告"

                

                # 表头样式

                hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

                hfont = Font(bold=True, color='FFFFFF')

                for col, h in enumerate(headers, 1):

                    cell = ws.cell(row=1, column=col, value=h)

                    cell.fill = hf

                    cell.font = hfont

                    cell.alignment = Alignment(horizontal='center')

                

                # 数据

                for row_idx, item in enumerate(history, 2):

                    action_tag = "减肥" if item.get("action") == "slim" else "脱敏"

                    orig_sz_col, new_sz_col = get_size_cols(item)

                    ws.cell(row=row_idx, column=1, value=item.get("time", ""))

                    ws.cell(row=row_idx, column=2, value=item.get("file_name", ""))

                    ws.cell(row=row_idx, column=3, value=action_tag)

                    ws.cell(row=row_idx, column=4, value=orig_sz_col)

                    ws.cell(row=row_idx, column=5, value=new_sz_col)

                    ws.cell(row=row_idx, column=6, value=get_effect(item))





                # 列宽

                ws.column_dimensions['A'].width = 20

                ws.column_dimensions['B'].width = 40

                ws.column_dimensions['C'].width = 10

                ws.column_dimensions['D'].width = 18

                ws.column_dimensions['E'].width = 18

                ws.column_dimensions['F'].width = 22

                ws.column_dimensions['G'].width = 30

                

                wb.save(file_path)

            

            from PySide6.QtWidgets import QMessageBox

            QMessageBox.information(self, "完成", f"报告已导出到：\n{file_path}")

        except Exception as e:

            from PySide6.QtWidgets import QMessageBox

            QMessageBox.critical(self, "错误", f"导出失败：{e}")





    def _format_size(self, size):

        """格式化文件大小"""

        try:

            size = int(size)

            if size < 1024:

                return f"{size} B"

            elif size < 1024 * 1024:

                return f"{size/1024:.1f} KB"

            else:

                return f"{size/(1024*1024):.2f} MB"

        except:

            return "0 B"



    def add_record(self, file_name, file_path, action, original_size, new_size, output_path=None):

        """添加记录（供其他标签页调用）"""

        return self.history_manager.add_record(

            file_name, file_path, action, original_size, new_size, output_path

        )

    

    def add_batch_record(self, total_files, success_count, total_saved, action='slim'):

        """添加批量处理记录"""

        return self.history_manager.add_batch_record(

            total_files, success_count, total_saved, action

        )

    

    def update_language(self, lang):

        """更新语言"""

        from translations import get_translation

        _ = lambda t: get_translation(t, lang)

        

        # 更新按钮

        if hasattr(self, 'btn_clear'):

            self.btn_clear.setText(_('清空历史'))

        if hasattr(self, 'btn_export'):

            self.btn_export.setText(_('导出历史'))

        

        # 更新表格标题

        if hasattr(self, 'table'):

            self.table.setHorizontalHeaderLabels([

                _('时间'), _('文件名'), _('操作'), _('原大小'), _('新大小'), _('状态'), _('详情')

            ])